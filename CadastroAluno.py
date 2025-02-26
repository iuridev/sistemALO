import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from datetime import datetime


class CadastroAluno:
    def __init__(self, root):
        self.root = root
        self.root.title("Cadastro de Aluno")

        # Labels e entradas para os dados do aluno
        ttk.Label(root, text="Nome:").grid(row=0, column=0, padx=5, pady=5)
        self.nome_entry = ttk.Entry(root)
        self.nome_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(root, text="RA:").grid(row=1, column=0, padx=5, pady=5)
        self.ra_entry = ttk.Entry(root)
        self.ra_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(root, text="Registro de Matrícula:").grid(
            row=2, column=0, padx=5, pady=5)
        self.registro_entry = ttk.Entry(root)
        self.registro_entry.grid(row=2, column=1, padx=5, pady=5)

        # Botão de cadastro
        ttk.Button(root, text="Cadastrar", command=self.cadastrar_aluno).grid(
            row=3, column=0, columnspan=2, pady=10)

    def cadastrar_aluno(self):
        nome = self.nome_entry.get()
        ra = self.ra_entry.get()
        registro = self.registro_entry.get()
        data_cadastro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not nome or not ra or not registro:
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return

        try:
            workbook = openpyxl.load_workbook("database\\cadasdroAlunos.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                ["ID", "Nome", "RA", "Registro de Matrícula", "Data de Cadastro"])

        id_aluno = sheet.max_row
        sheet.append([id_aluno, nome, ra, registro, data_cadastro])
        workbook.save("database\\cadasdroAlunos.xlsx")

        messagebox.showinfo("Sucesso", "Aluno cadastrado com sucesso!")
        self.nome_entry.delete(0, tk.END)
        self.ra_entry.delete(0, tk.END)
        self.registro_entry.delete(0, tk.END)


# Exemplo de uso
if __name__ == "__main__":
    root = tk.Tk()
    app = CadastroAluno(root)
    root.mainloop()
