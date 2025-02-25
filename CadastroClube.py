import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from datetime import datetime


class CadastroClube:
    def __init__(self, root):
        self.root = root
        self.root.title("Cadastro de Clube")

        # Labels e entradas para os dados do aluno
        ttk.Label(root, text="Nome:").grid(row=0, column=0, padx=5, pady=5)
        self.nome_entry = ttk.Entry(root)
        self.nome_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(root, text="Séries:").grid(row=1, column=0, padx=5, pady=5)
        self.series_entry = ttk.Entry(root)
        self.series_entry.grid(row=1, column=1, padx=5, pady=5)

        # Botão de cadastro
        ttk.Button(root, text="Cadastrar", command=self.cadastrar_clube).grid(
            row=2, column=0, columnspan=2, pady=10)

    def cadastrar_clube(self):
        nome = self.nome_entry.get()
        series = self.series_entry.get()
        data_cadastro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not nome or not series:
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return

        try:
            workbook = openpyxl.load_workbook("clubes.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                ["ID", "Nome", "Séries", "Data de Cadastro"])

        id_clube = sheet.max_row
        sheet.append([id_clube, nome, series, data_cadastro])
        workbook.save("clubes.xlsx")

        messagebox.showinfo("Sucesso", "Clube cadastrado com sucesso!")
        self.nome_entry.delete(0, tk.END)
        self.series_entry.delete(0, tk.END)


# Exemplo de uso
if __name__ == "__main__":
    root = tk.Tk()
    app = CadastroClube(root)
    root.mainloop()
