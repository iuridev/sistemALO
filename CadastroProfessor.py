import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from datetime import datetime


class CadastroProfessor:
    def __init__(self, root):
        self.root = root
        self.root.title("Cadastro de Professor")

        # Labels e entradas para os dados do aluno
        ttk.Label(root, text="Nome:").grid(row=0, column=0, padx=5, pady=5)
        self.nome_entry = ttk.Entry(root)
        self.nome_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(root, text="RG:").grid(row=1, column=0, padx=5, pady=5)
        self.rg_entry = ttk.Entry(root)
        self.rg_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(root, text="CPF:").grid(
            row=2, column=0, padx=5, pady=5)
        self.cpf_entry = ttk.Entry(root)
        self.cpf_entry.grid(row=2, column=1, padx=5, pady=5)

        # Botão de cadastro
        ttk.Button(root, text="Cadastrar", command=self.cadastrar_professor).grid(
            row=3, column=0, columnspan=2, pady=10)

    def cadastrar_professor(self):
        nome = self.nome_entry.get()
        rg = self.rg_entry.get()
        cpf = self.cpf_entry.get()
        data_cadastro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not nome or not rg or not cpf:
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return

        try:
            workbook = openpyxl.load_workbook("database\\professores.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                ["ID", "Nome", "RG", "CPF", "Data de Cadastro"])

        id_professor = sheet.max_row
        sheet.append([id_professor, nome, rg, cpf, data_cadastro])
        workbook.save("database\\professores.xlsx")

        messagebox.showinfo("Sucesso", "Professor cadastrado com sucesso!")
        self.nome_entry.delete(0, tk.END)
        self.rg_entry.delete(0, tk.END)
        self.cpf_entry.delete(0, tk.END)


# Exemplo de uso
if __name__ == "__main__":
    root = tk.Tk()
    app = CadastroProfessor(root)
    root.mainloop()
