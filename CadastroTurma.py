import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from datetime import datetime


class CadastroTurma:
    def __init__(self, root):
        self.root = root
        self.root.title("Cadastro de Turma")

        # Labels e entradas para os dados do aluno
        ttk.Label(root, text="Série:").grid(row=0, column=0, padx=5, pady=5)
        self.serie_entry = ttk.Entry(root)
        self.serie_entry.grid(row=0, column=1, padx=5, pady=5)

        ttk.Label(root, text="Turma:").grid(row=1, column=0, padx=5, pady=5)
        self.turma_entry = ttk.Entry(root)
        self.turma_entry.grid(row=1, column=1, padx=5, pady=5)

        ttk.Label(root, text="Tipo Ensino:").grid(
            row=2, column=0, padx=5, pady=5)
        self.tipo_entry = ttk.Entry(root)
        self.tipo_entry.grid(row=2, column=1, padx=5, pady=5)

        # Botão de cadastro
        ttk.Button(root, text="Cadastrar", command=self.cadastrar_turma).grid(
            row=3, column=0, columnspan=2, pady=10)

    def cadastrar_turma(self):
        serie = self.serie_entry.get()
        turma = self.turma_entry.get()
        tipo = self.tipo_entry.get()

        data_cadastro = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not serie or not turma or not tipo:
            messagebox.showerror("Erro", "Preencha todos os campos!")
            return

        try:
            workbook = openpyxl.load_workbook("database\\cadasdroTurmas.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                ["ID", "Séries", "Turma", "Tipo", "Data de Cadastro"])

        id_turma = sheet.max_row
        sheet.append([id_turma, serie, turma, tipo, data_cadastro])
        workbook.save("database\\cadasdroTurmas.xlsx")

        messagebox.showinfo("Sucesso", "Turma cadastrado com sucesso!")
        self.serie_entry.delete(0, tk.END)
        self.turma_entry.delete(0, tk.END)
        self.tipo_entry.delete(0, tk.END)


# Exemplo de uso
if __name__ == "__main__":
    root = tk.Tk()
    app = CadastroTurma(root)
    root.mainloop()
